# -*- coding: utf-8 -*-
import logging
import os
import sys
from collections import OrderedDict, Counter
from functools import reduce

import numpy as np
import torch
from torch import optim, nn
from torch.utils.data import DataLoader, TensorDataset, ConcatDataset

from tqdm import tqdm
import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

from sklearn.metrics import cohen_kappa_score, f1_score, roc_auc_score

# ===================== 这里把网络换成 ADFCNNNet =====================
# 例如：from Net.ADFCNN import ADFCNNNet
from Net.ADFCNN import ADFCNNNet  # 

from Net.MaxNorm import MaxNormDefaultConstraint

from braindecode.datasets.bcic_iv_2a import BCICompetition4Set2A
from braindecode.mne_ext.signalproc import mne_apply
from braindecode.datautil.signalproc import bandpass_cnt, exponential_running_standardize
from braindecode.datautil.trial_segment import create_signal_target_from_raw_mne
from braindecode.datautil.signal_target import SignalAndTarget
from braindecode.torch_ext.modules import Expression


# ===================== 关键：为了“不改你后面stage1/2/3逻辑”，做一个薄封装，让 ADFCNN 也有 backbone/sequential/classifier =====================
def _to_adfcnn_input(x):
    """
    你的数据： (B, C, T, 1)
    ADFCNN期望：(B, 1, C, T)
    """
    if x.ndim != 4:
        raise ValueError(f"Unexpected input ndim={x.ndim}, expect 4D (B,C,T,1). Got {x.shape}")
    if x.shape[-1] != 1:
        raise ValueError(f"Unexpected last dim, expect 1. Got {x.shape}")
    return x.permute(0, 3, 1, 2).contiguous()  # (B,1,C,T)


def _feat_to_disc_shape(feat):
    """
    ADFCNN backbone 输出通常是 (B, F, 1, W)
    判别器期望是 (B, F, T_len, 1)（你原来的 _squeeze_final_output 假设 dim3==1）
    因此交换最后两维： (B, F, 1, W) -> (B, F, W, 1)
    """
    if feat.ndim != 4:
        raise ValueError(f"Unexpected feat ndim={feat.ndim}. Got {feat.shape}")
    if feat.shape[2] == 1:
        return feat.permute(0, 1, 3, 2).contiguous()  # (B,F,W,1)
    # 若你实现里不是 (B,F,1,W)，那就尽量保证最后一维为1
    if feat.shape[-1] == 1:
        return feat
    raise ValueError(f"Unexpected ADFCNN feat shape, cannot adapt: {feat.shape}")


class ADFCNNBackbone(nn.Module):
    """只做输入/输出维度适配，复用你已有的 ADFCNNNet.backbone"""
    def __init__(self, adfnet: ADFCNNNet):
        super().__init__()
        self.net_backbone = adfnet.backbone

    def forward(self, x):
        x = _to_adfcnn_input(x)         # (B,1,C,T)
        feat = self.net_backbone(x)     # (B,F,1,W)  (一般如此)
        feat = _feat_to_disc_shape(feat)  # -> (B,F,W,1)
        return feat


class ADFCNNNetWrapper(nn.Module):
    """
    让 ADFCNNNet 兼容你现有代码：
      - model.backbone(data) -> (B, F, T_len, 1)
      - model.sequential(...) -> Identity（保持不变）
      - model.classifier(feature) -> log-prob (B, n_classes)
      - model(data) -> log-prob (B, n_classes)
    """
    def __init__(self, n_classes=4, in_chans=22, in_samples=1125, return_fc=False):
        super().__init__()
        # 注意：你给的 ADFCNNNet 用 sampling_rate 来判断 linear_in，
        # 这里直接把 in_samples(=1125) 传进去即可匹配它的 if/elif 逻辑。
        adf = ADFCNNNet(
            num_classes=n_classes,
            num_channels=in_chans,
            sampling_rate=in_samples,
            return_fc=return_fc
        )
        self.backbone = ADFCNNBackbone(adf)
        self.sequential = nn.Identity()
        self.classifier = adf.classifier  # 直接复用，内部 flatten + Linear + LogSoftmax

    def forward(self, x):
        feat = self.sequential(self.backbone(x))
        out = self.classifier(feat)
        # 如果你把 return_fc=True，会返回 (fc_feat, logprob)，这里做一下兼容
        if isinstance(out, (tuple, list)) and len(out) == 2:
            _, out = out
        return out


# ===================== 全局配置（请根据你的环境修改） =====================
SAVE_ROOT = r".\ADFCNN0hz"
DATA_FOLDER = r"D:\BCI\BCICIV_2a_gdf"
RANDOM_SEED = 42

SHOT_TOTAL_LIST = [20, 40, 60, 80, 120, 160, 200, 240]
N_CLASSES = 4

BATCH_SIZE = 32
EPOCHS = 300
LR = 1e-3
LAMBDA1 = 1
LAMBDA2 = 0.4

stage1_metrics = []
stage2_metrics = []
stage3_metrics = []
final_results = {}

target_only_metrics = []
source_only_metrics = []
source_only_euclid_metrics = []

TARGET_SUBJECT = 1
SAMPLES_PER_CLASS = 60
save = SAVE_ROOT


def check_and_create_folder(folder_path: str):
    if not os.path.exists(folder_path):
        os.makedirs(folder_path, exist_ok=True)


def reset_global_metrics():
    global stage1_metrics, stage2_metrics, stage3_metrics, final_results
    global target_only_metrics, source_only_metrics, source_only_euclid_metrics
    stage1_metrics = []
    stage2_metrics = []
    stage3_metrics = []
    target_only_metrics = []
    source_only_metrics = []
    source_only_euclid_metrics = []
    final_results = {}


def calculate_accuracy(logits, labels):
    preds = torch.argmax(logits, dim=1)
    return (preds == labels).float().mean().item()


def freeze_model(model):
    for p in model.parameters():
        p.requires_grad = False


def unfreeze_model(model):
    for p in model.parameters():
        p.requires_grad = True


def subject_adv_loss(predictions, target_prob=1 / 9):
    predictions = torch.softmax(predictions, dim=1)
    loss_per_sample = (predictions - target_prob) ** 2
    return torch.mean(torch.sum(loss_per_sample, dim=1))


def domain_adv_loss(predictions, target_prob=0.5):
    predictions = torch.softmax(predictions, dim=1)
    loss_per_sample = (predictions - target_prob) ** 2
    return torch.mean(torch.sum(loss_per_sample, dim=1))


def _safe_auc(y_true: np.ndarray, y_prob: np.ndarray, n_classes: int) -> float:
    try:
        return float(roc_auc_score(y_true, y_prob, multi_class="ovr", average="macro"))
    except Exception:
        try:
            y_true_oh = np.eye(n_classes)[y_true]
            return float(roc_auc_score(y_true_oh, y_prob, average="macro"))
        except Exception:
            return float("nan")


def eval_model_metrics(model, loader, device, n_classes: int):
    model.eval()
    all_probs, all_preds, all_labels = [], [], []

    with torch.no_grad():
        for batch in loader:
            x = batch[0].to(device)
            y = batch[1].to(device)

            out = model(x)
            prob = torch.softmax(out, dim=1)
            pred = torch.argmax(prob, dim=1)

            all_probs.append(prob.detach().cpu().numpy())
            all_preds.append(pred.detach().cpu().numpy())
            all_labels.append(y.detach().cpu().numpy())

    y_true = np.concatenate(all_labels, axis=0).astype(int)
    y_pred = np.concatenate(all_preds, axis=0).astype(int)
    y_prob = np.concatenate(all_probs, axis=0)

    acc = float((y_pred == y_true).mean())
    kappa = float(cohen_kappa_score(y_true, y_pred))
    f1 = float(f1_score(y_true, y_pred, average="macro"))
    auc = _safe_auc(y_true, y_prob, n_classes=n_classes)

    return {"acc": acc, "kappa": kappa, "f1": f1, "auc": auc}


# ===================== 欧式对齐（完全使用你给的版本，不改动） =====================
def euclidAlign(X):
    R = np.zeros((X.shape[1], X.shape[1]))
    R = reduce(lambda R, y: R + np.dot(y, y.T), X, R)
    R = R / X.shape[0]
    for i in range(X.shape[0]):
        X[i, :, :] = np.dot(fuduiban(R), X[i, :, :])
    return X


def fuduiban(R):
    v, Q = np.linalg.eig(R)
    ss = np.diag(v ** (-0.5))
    ss[np.isnan(ss)] = 0
    re = np.dot(Q, np.dot(ss, np.linalg.inv(Q)))
    return np.real(re)


def apply_euclid_align_to_4d(X4: np.ndarray) -> np.ndarray:
    if X4.ndim == 4:
        assert X4.shape[-1] == 1
        X3 = X4[..., 0]
    elif X4.ndim == 3:
        X3 = X4
    else:
        raise ValueError(f"Unexpected X shape: {X4.shape}")

    X3_aligned = euclidAlign(X3.copy())
    X4_aligned = np.expand_dims(X3_aligned, axis=3)
    return X4_aligned.astype(np.float32)


# ===================== 数据处理函数 =====================
def _normalize_labels_to_zero_based(y: np.ndarray, n_classes=4):
    y = y.astype(np.int64)
    if y.min() == 1 and y.max() == n_classes:
        y = y - 1
    return y


def preprocessing_2a(subject_id, data_folder):
    low_cut_hz = 0
    high_cut_hz = 38
    factor_new = 1e-3
    init_block_size = 1000
    ival = [-500, 4000]

    train_filename = "A{:02d}T.gdf".format(subject_id)
    test_filename = "A{:02d}E.gdf".format(subject_id)
    train_filepath = os.path.join(data_folder, train_filename)
    test_filepath = os.path.join(data_folder, test_filename)
    train_label_filepath = train_filepath.replace(".gdf", ".mat")
    test_label_filepath = test_filepath.replace(".gdf", ".mat")

    train_loader = BCICompetition4Set2A(train_filepath, labels_filename=train_label_filepath)
    test_loader = BCICompetition4Set2A(test_filepath, labels_filename=test_label_filepath)
    train_cnt = train_loader.load()
    test_cnt = test_loader.load()

    train_cnt = train_cnt.drop_channels(["STI 014", "EOG-left", "EOG-central", "EOG-right"])
    assert len(train_cnt.ch_names) == 22
    train_cnt = mne_apply(lambda a: a * 1e6, train_cnt)
    train_cnt = mne_apply(
        lambda a: bandpass_cnt(a, low_cut_hz, high_cut_hz, train_cnt.info["sfreq"], filt_order=3, axis=1), train_cnt
    )
    train_cnt = mne_apply(
        lambda a: exponential_running_standardize(
            a.T, factor_new=factor_new, init_block_size=init_block_size, eps=1e-4
        ).T,
        train_cnt,
    )

    test_cnt = test_cnt.drop_channels(["EOG-central", "EOG-right", "STI 014", "EOG-left"])
    assert len(test_cnt.ch_names) == 22
    test_cnt = mne_apply(lambda a: a * 1e6, test_cnt)
    test_cnt = mne_apply(
        lambda a: bandpass_cnt(a, low_cut_hz, high_cut_hz, test_cnt.info["sfreq"], filt_order=3, axis=1), test_cnt
    )
    test_cnt = mne_apply(
        lambda a: exponential_running_standardize(
            a.T, factor_new=factor_new, init_block_size=init_block_size, eps=1e-4
        ).T,
        test_cnt,
    )

    marker_def = OrderedDict([("Left Hand", [1]), ("Right Hand", [2]), ("Foot", [3]), ("Tongue", [4])])
    train_set = create_signal_target_from_raw_mne(train_cnt, marker_def, ival)
    test_set = create_signal_target_from_raw_mne(test_cnt, marker_def, ival)

    train_set.y = _normalize_labels_to_zero_based(train_set.y, n_classes=4)
    test_set.y = _normalize_labels_to_zero_based(test_set.y, n_classes=4)

    return train_set, test_set


def sample_k_from_target_test(target_test_set_full, samples_per_class, random_seed=RANDOM_SEED):
    np.random.seed(random_seed)
    torch.manual_seed(random_seed)

    X = target_test_set_full.X
    y = target_test_set_full.y

    unique_labels = sorted(np.unique(y))
    print(f"\n目标域sub{TARGET_SUBJECT}全量测试数据标签分布：{Counter(y)}")
    assert len(unique_labels) == N_CLASSES, f"预期{N_CLASSES}类，实际{len(unique_labels)}类：{unique_labels}"

    train_indices = []
    for cls in unique_labels:
        cls_idx = np.where(y == cls)[0]
        assert len(cls_idx) >= samples_per_class, f"类别{cls}样本不足{samples_per_class}（仅{len(cls_idx)}）"
        sampled_idx = np.random.choice(cls_idx, size=samples_per_class, replace=False)
        train_indices.extend(sampled_idx.tolist())

    train_indices = np.array(train_indices, dtype=int)
    all_indices = np.arange(len(X))
    val_indices = np.setdiff1d(all_indices, train_indices)

    train_from_test_X = X[train_indices]
    train_from_test_y = y[train_indices]
    val_from_test_X = X[val_indices]
    val_from_test_y = y[val_indices]

    print(f"抽取few-shot样本标签分布：{Counter(train_from_test_y)}")
    print(f"从目标域sub{TARGET_SUBJECT}测试集抽取：{len(train_indices)}个（每类{samples_per_class}个）")
    print(f"目标域sub{TARGET_SUBJECT}验证集：{len(val_indices)}个，标签分布：{Counter(val_from_test_y)}")

    target_shot_set = SignalAndTarget(train_from_test_X, train_from_test_y)
    target_val_set = SignalAndTarget(val_from_test_X, val_from_test_y)
    return target_shot_set, target_val_set, train_indices


def load_all_data(data_folder, target_subject, batch_size, samples_per_class):
    sub_list = [1, 2, 3, 4, 5, 6, 7, 8, 9]
    all_train_sets, all_test_sets = [], []

    for sub_id in sub_list:
        train_set, test_set = preprocessing_2a(sub_id, data_folder)
        train_set.X = np.expand_dims(train_set.X, axis=3).astype(np.float32)
        train_set.y = train_set.y.astype(np.int64)
        test_set.X = np.expand_dims(test_set.X, axis=3).astype(np.float32)
        test_set.y = test_set.y.astype(np.int64)
        all_train_sets.append(train_set)
        all_test_sets.append(test_set)

    target_train_set = all_train_sets[target_subject - 1]
    target_test_set_full = all_test_sets[target_subject - 1]

    target_shot_set, target_val_set, train_indices = sample_k_from_target_test(
        target_test_set_full, samples_per_class=samples_per_class
    )

    source_train_sets = [(sub_id, all_train_sets[sub_id - 1]) for sub_id in sub_list if sub_id != target_subject]
    source_train_X = np.concatenate([s.X for (_, s) in source_train_sets], axis=0)
    source_train_y = np.concatenate([s.y for (_, s) in source_train_sets], axis=0)
    source_train_set = SignalAndTarget(source_train_X, source_train_y)
    print(f"\n源域sub!=target全量训练集：{len(source_train_X)}个样本")

    combined_train_X = np.concatenate([source_train_X, target_shot_set.X], axis=0)
    combined_train_y = np.concatenate([source_train_y, target_shot_set.y], axis=0)
    combined_train_set = SignalAndTarget(combined_train_X, combined_train_y)

    source_sample_num = len(source_train_X)
    print(f"\n合并后总训练数据：{combined_train_X.shape[0]}（源域{source_sample_num} + 目标few-shot{len(target_shot_set.X)}）")

    domain_labels = np.zeros(len(combined_train_X), dtype=np.int64)
    domain_labels[source_sample_num:] = 1

    class TrainDataset(TensorDataset):
        def __init__(self, signal_target, domain_labels_):
            self.X = torch.from_numpy(signal_target.X).float()
            self.y = torch.from_numpy(signal_target.y).long()
            self.domain_labels = torch.from_numpy(domain_labels_).long()
            super().__init__(self.X, self.y, self.domain_labels)

        def __getitem__(self, idx):
            return self.X[idx], self.y[idx], self.domain_labels[idx]

    class Stage2Dataset(TensorDataset):
        def __init__(self, signal_target, domain_label, subject_label):
            self.X = torch.from_numpy(signal_target.X).float()
            self.y = torch.from_numpy(signal_target.y).long()
            self.domain_label = domain_label
            self.subject_label = subject_label
            super().__init__(self.X, self.y)

        def __getitem__(self, idx):
            domain_tensor = torch.tensor(self.domain_label, dtype=torch.long)
            subject_tensor = torch.tensor(self.subject_label, dtype=torch.long)
            return self.X[idx], self.y[idx], domain_tensor, subject_tensor

    train_loader = DataLoader(TrainDataset(combined_train_set, domain_labels), batch_size=batch_size, shuffle=True, num_workers=0)
    target_shot_loader = DataLoader(TensorDataset(torch.from_numpy(target_shot_set.X).float(), torch.from_numpy(target_shot_set.y).long()),
                                    batch_size=batch_size, shuffle=False, num_workers=0)
    target_val_loader = DataLoader(TensorDataset(torch.from_numpy(target_val_set.X).float(), torch.from_numpy(target_val_set.y).long()),
                                   batch_size=batch_size, shuffle=False, num_workers=0)
    target_full_test_loader = DataLoader(TensorDataset(torch.from_numpy(target_test_set_full.X).float(), torch.from_numpy(target_test_set_full.y).long()),
                                         batch_size=batch_size, shuffle=False, num_workers=0)
    target_train_loader = DataLoader(TensorDataset(torch.from_numpy(target_train_set.X).float(), torch.from_numpy(target_train_set.y).long()),
                                     batch_size=batch_size, shuffle=True, num_workers=0)
    source_train_loader = DataLoader(TensorDataset(torch.from_numpy(source_train_set.X).float(), torch.from_numpy(source_train_set.y).long()),
                                     batch_size=batch_size, shuffle=True, num_workers=0)

    print("\n========== 对比方法3：构造欧式对齐后的训练/验证/测试数据 ==========")

    source_train_X_euclid = apply_euclid_align_to_4d(source_train_set.X)
    source_train_set_euclid = SignalAndTarget(source_train_X_euclid, source_train_set.y.copy())
    source_train_loader_euclid = DataLoader(
        TensorDataset(torch.from_numpy(source_train_set_euclid.X).float(), torch.from_numpy(source_train_set_euclid.y).long()),
        batch_size=batch_size, shuffle=True, num_workers=0
    )

    target_shot_X_euclid = apply_euclid_align_to_4d(target_shot_set.X)
    target_shot_set_euclid = SignalAndTarget(target_shot_X_euclid, target_shot_set.y.copy())
    target_shot_loader_euclid = DataLoader(
        TensorDataset(torch.from_numpy(target_shot_set_euclid.X).float(), torch.from_numpy(target_shot_set_euclid.y).long()),
        batch_size=batch_size, shuffle=False, num_workers=0
    )

    target_val_X_euclid = apply_euclid_align_to_4d(target_val_set.X)
    target_val_set_euclid = SignalAndTarget(target_val_X_euclid, target_val_set.y.copy())
    target_val_loader_euclid = DataLoader(
        TensorDataset(torch.from_numpy(target_val_set_euclid.X).float(), torch.from_numpy(target_val_set_euclid.y).long()),
        batch_size=batch_size, shuffle=False, num_workers=0
    )

    target_full_X_euclid = apply_euclid_align_to_4d(target_test_set_full.X)
    target_full_set_euclid = SignalAndTarget(target_full_X_euclid, target_test_set_full.y.copy())
    target_full_test_loader_euclid = DataLoader(
        TensorDataset(torch.from_numpy(target_full_set_euclid.X).float(), torch.from_numpy(target_full_set_euclid.y).long()),
        batch_size=batch_size, shuffle=False, num_workers=0
    )

    print(f"source_train_euclid: {source_train_set_euclid.X.shape}")
    print(f"target_shot_euclid : {target_shot_set_euclid.X.shape}")
    print(f"target_val_euclid  : {target_val_set_euclid.X.shape}")
    print(f"target_full_euclid : {target_full_set_euclid.X.shape}")

    stage2_train_list = []
    for sub_id, s in source_train_sets:
        stage2_train_list.append(Stage2Dataset(s, domain_label=0, subject_label=sub_id - 1))
    stage2_train_list.append(Stage2Dataset(target_shot_set, domain_label=1, subject_label=target_subject - 1))
    stage2_train_loader = DataLoader(ConcatDataset(stage2_train_list), batch_size=batch_size, shuffle=True, num_workers=0)

    return (
        train_loader, stage2_train_loader, target_train_loader, source_train_loader,
        target_shot_loader, target_val_loader, target_full_test_loader,
        source_train_loader_euclid, target_shot_loader_euclid, target_val_loader_euclid, target_full_test_loader_euclid,
        source_sample_num, target_train_set, train_indices
    )


# ===================== 判别器定义 =====================
def _squeeze_final_output(x):
    assert x.size()[3] == 1
    x = x[:, :, :, 0]
    if x.size()[2] == 1:
        x = x[:, :, 0]
    return x


class ClassifierLikeDiscriminator(nn.Module):
    def __init__(self, in_channels, in_length, out_classes):
        super().__init__()
        self.discriminator = nn.Sequential(
            nn.Conv2d(in_channels, out_classes, (in_length, 1), bias=True),
            nn.LogSoftmax(dim=1),
            Expression(_squeeze_final_output),
        )
        self._initialize_weights()

    def _initialize_weights(self):
        conv_layer = self.discriminator[0]
        nn.init.xavier_uniform_(conv_layer.weight, gain=1)
        nn.init.constant_(conv_layer.bias, 0)

    def forward(self, x):
        return self.discriminator(x)


def create_domain_discriminator(in_channels, in_length):
    return ClassifierLikeDiscriminator(in_channels, in_length, out_classes=2)


def create_subject_discriminator(in_channels, in_length):
    return ClassifierLikeDiscriminator(in_channels, in_length, out_classes=9)


# ===================== Stage1/2/3 & baselines（保持你原逻辑） =====================
def stage1_unsupervised_train(model, train_loader, val_loader, test_loader, full_test_loader, device,
                              epochs=300, lr=5e-4, n_classes=4):
    global stage1_metrics
    model.to(device)

    criterion = nn.NLLLoss()
    optimizer = optim.Adam(model.parameters(), lr=lr)
    constraint = MaxNormDefaultConstraint()

    best_val_metrics = None
    best_test_metrics = None
    best_full_metrics = None

    for epoch in range(epochs):
        model.train()
        train_loss = 0.0
        train_acc = 0.0
        source_sample_count = 0

        pbar = tqdm(enumerate(train_loader), total=len(train_loader), leave=False)
        for _, (data, label, domain_label) in pbar:
            data, label = data.to(device), label.to(device)
            optimizer.zero_grad()
            output = model(data)

            is_source = (domain_label == 0).to(device)
            if torch.any(is_source):
                source_output = output[is_source]
                source_label = label[is_source]
                class_loss = criterion(source_output, source_label)

                train_loss += class_loss.item() * source_output.shape[0]
                train_acc += calculate_accuracy(source_output, source_label) * source_output.shape[0]
                source_sample_count += source_output.shape[0]

                class_loss.backward()
                optimizer.step()
                constraint.apply(model)

        train_loss /= max(source_sample_count, 1)
        train_acc /= max(source_sample_count, 1)

        val_m = eval_model_metrics(model, val_loader, device, n_classes=n_classes)
        test_m = eval_model_metrics(model, test_loader, device, n_classes=n_classes)
        full_m = eval_model_metrics(model, full_test_loader, device, n_classes=n_classes)

        if (best_val_metrics is None) or (val_m["acc"] > best_val_metrics["acc"]):
            best_val_metrics = val_m
            best_test_metrics = test_m
            best_full_metrics = full_m
            torch.save(
                {
                    "backbone": model.backbone.state_dict(),
                    "sequential": model.sequential.state_dict(),
                    "classifier": model.classifier.state_dict(),
                },
                f"{save}/stage1_best.pth",
            )

        stage1_metrics.append({
            "Epoch": epoch + 1,
            "Train_Loss": round(train_loss, 4),
            "Train_Acc": round(train_acc, 4),

            "Val_Acc(acc2)": round(val_m["acc"], 4),
            "Val_Kappa": round(val_m["kappa"], 4),
            "Val_F1": round(val_m["f1"], 4),
            "Val_AUC": round(val_m["auc"], 4) if not np.isnan(val_m["auc"]) else val_m["auc"],

            "Test_Acc(acc1)": round(test_m["acc"], 4),
            "Test_Kappa": round(test_m["kappa"], 4),
            "Test_F1": round(test_m["f1"], 4),
            "Test_AUC": round(test_m["auc"], 4) if not np.isnan(test_m["auc"]) else test_m["auc"],

            "Full_Acc(acc3)": round(full_m["acc"], 4),
            "Full_Kappa": round(full_m["kappa"], 4),
            "Full_F1": round(full_m["f1"], 4),
            "Full_AUC": round(full_m["auc"], 4) if not np.isnan(full_m["auc"]) else full_m["auc"],

            "Best_Val_Acc": round(best_val_metrics["acc"], 4) if best_val_metrics else 0.0,
        })

    checkpoint = torch.load(f"{save}/stage1_best.pth", map_location=device)
    model.backbone.load_state_dict(checkpoint["backbone"])
    model.sequential.load_state_dict(checkpoint["sequential"])
    model.classifier.load_state_dict(checkpoint["classifier"])

    return model, best_val_metrics, best_test_metrics, best_full_metrics


def stage2_train_discriminators(backbone, sequential, domain_disc, subject_disc,
                                train_loader, val_loader, test_loader, full_test_loader,
                                device, target_subject_idx, epochs=300, lr=5e-4):
    global stage2_metrics
    freeze_model(backbone)
    freeze_model(sequential)
    backbone.to(device).eval()
    sequential.to(device).eval()
    domain_disc.to(device)
    subject_disc.to(device)

    criterion = nn.NLLLoss()
    opt_domain = optim.Adam(domain_disc.parameters(), lr=lr)
    opt_subject = optim.Adam(subject_disc.parameters(), lr=lr)
    constraint = MaxNormDefaultConstraint()

    best_val_domain_acc = 0.0
    best_val_subject_acc = 0.0

    best_test_domain_acc = 0.0
    best_test_subject_acc = 0.0
    best_full_domain_acc = 0.0
    best_full_subject_acc = 0.0

    for epoch in range(epochs):
        domain_disc.train()
        subject_disc.train()

        train_domain_loss = 0.0
        train_subject_loss = 0.0
        train_domain_acc = 0.0
        train_subject_acc = 0.0

        pbar = tqdm(enumerate(train_loader), total=len(train_loader), leave=False)
        for _, (data, _, domain_label, subject_label) in pbar:
            data = data.to(device)
            domain_label = domain_label.to(device)
            subject_label = subject_label.to(device)

            with torch.no_grad():
                features = sequential(backbone(data))

            opt_domain.zero_grad()
            d_logits = domain_disc(features)
            d_loss = criterion(d_logits, domain_label)
            d_loss.backward()
            opt_domain.step()
            constraint.apply(domain_disc)

            opt_subject.zero_grad()
            s_logits = subject_disc(features)
            s_loss = criterion(s_logits, subject_label)
            s_loss.backward()
            opt_subject.step()
            constraint.apply(subject_disc)

            bs = data.size(0)
            train_domain_loss += d_loss.item() * bs
            train_subject_loss += s_loss.item() * bs
            train_domain_acc += calculate_accuracy(d_logits, domain_label) * bs
            train_subject_acc += calculate_accuracy(s_logits, subject_label) * bs

        train_domain_loss /= len(train_loader.dataset)
        train_subject_loss /= len(train_loader.dataset)
        train_domain_acc /= len(train_loader.dataset)
        train_subject_acc /= len(train_loader.dataset)

        def _eval_disc_acc(loader):
            domain_disc.eval()
            subject_disc.eval()
            d_acc = 0.0
            s_acc = 0.0
            with torch.no_grad():
                for data, _ in loader:
                    data = data.to(device)
                    domain_true = torch.ones(len(data), dtype=torch.long, device=device)
                    subject_true = torch.full((len(data),), target_subject_idx, dtype=torch.long, device=device)

                    feat = sequential(backbone(data))
                    d_logits = domain_disc(feat)
                    s_logits = subject_disc(feat)

                    d_acc += calculate_accuracy(d_logits, domain_true) * len(data)
                    s_acc += calculate_accuracy(s_logits, subject_true) * len(data)

            d_acc /= len(loader.dataset)
            s_acc /= len(loader.dataset)
            return d_acc, s_acc

        val_d_acc, val_s_acc = _eval_disc_acc(val_loader)
        test_d_acc, test_s_acc = _eval_disc_acc(test_loader)
        full_d_acc, full_s_acc = _eval_disc_acc(full_test_loader)

        if val_d_acc > best_val_domain_acc:
            best_val_domain_acc = val_d_acc
            best_test_domain_acc = test_d_acc
            best_full_domain_acc = full_d_acc
            torch.save(domain_disc.state_dict(), f"{save}/domain_disc_best.pth")

        if val_s_acc > best_val_subject_acc:
            best_val_subject_acc = val_s_acc
            best_test_subject_acc = test_s_acc
            best_full_subject_acc = full_s_acc
            torch.save(subject_disc.state_dict(), f"{save}/subject_disc_best.pth")

        stage2_metrics.append({
            "Epoch": epoch + 1,
            "Domain_Train_Loss": round(train_domain_loss, 4),
            "Domain_Train_Acc": round(train_domain_acc, 4),
            "Subject_Train_Loss": round(train_subject_loss, 4),
            "Subject_Train_Acc": round(train_subject_acc, 4),

            "Val_Domain_Acc(acc2)": round(val_d_acc, 4),
            "Val_Subject_Acc(acc2)": round(val_s_acc, 4),
            "Test_Domain_Acc(acc1)": round(test_d_acc, 4),
            "Test_Subject_Acc(acc1)": round(test_s_acc, 4),
            "Full_Domain_Acc(acc3)": round(full_d_acc, 4),
            "Full_Subject_Acc(acc3)": round(full_s_acc, 4),

            "Best_Val_Domain_Acc": round(best_val_domain_acc, 4),
            "Best_Val_Subject_Acc": round(best_val_subject_acc, 4),
        })

    domain_disc.load_state_dict(torch.load(f"{save}/domain_disc_best.pth", map_location=device))
    subject_disc.load_state_dict(torch.load(f"{save}/subject_disc_best.pth", map_location=device))

    return domain_disc, subject_disc, best_test_domain_acc, best_test_subject_acc, best_full_domain_acc, best_full_subject_acc


def stage3_adversarial_train(model, domain_disc, subject_disc,
                             train_loader, val_loader, test_loader, full_test_loader,
                             device, epochs=300, lr=5e-4, lambda1=0.05, lambda2=0.05, n_classes=4):
    global stage3_metrics
    freeze_model(domain_disc)
    freeze_model(subject_disc)
    unfreeze_model(model.backbone)
    unfreeze_model(model.sequential)
    unfreeze_model(model.classifier)

    model.to(device)
    domain_disc.to(device).eval()
    subject_disc.to(device).eval()

    optimizer = optim.Adam(
        list(model.backbone.parameters()) + list(model.sequential.parameters()) + list(model.classifier.parameters()),
        lr=lr,
    )

    class_criterion = nn.NLLLoss()
    constraint = MaxNormDefaultConstraint()

    best_val_metrics = None
    best_test_metrics = None
    best_full_metrics = None

    for epoch in range(epochs):
        model.train()
        total_loss = 0.0
        class_loss_total = 0.0
        train_acc = 0.0
        source_sample_count = 0

        pbar = tqdm(enumerate(train_loader), total=len(train_loader), leave=False)
        for _, (data, label, domain_label) in pbar:
            data, label = data.to(device), label.to(device)
            optimizer.zero_grad()

            features = model.sequential(model.backbone(data))
            class_output = model.classifier(features)
            if isinstance(class_output, (tuple, list)) and len(class_output) == 2:
                _, class_output = class_output

            is_source = (domain_label == 0).to(device)
            class_loss = torch.tensor(0.0, device=device)
            if torch.any(is_source):
                source_out = class_output[is_source]
                source_label = label[is_source]
                class_loss = class_criterion(source_out, source_label)

                class_loss_total += class_loss.item() * source_out.shape[0]
                train_acc += calculate_accuracy(source_out, source_label) * source_out.shape[0]
                source_sample_count += source_out.shape[0]

            d_adv = domain_adv_loss(domain_disc(features))
            s_adv = subject_adv_loss(subject_disc(features))

            loss = class_loss + lambda1 * d_adv + lambda2 * s_adv
            loss.backward()
            optimizer.step()

            constraint.apply(model.backbone)
            constraint.apply(model.sequential)
            constraint.apply(model.classifier)

            bs = data.size(0)
            total_loss += loss.item() * bs

        total_loss /= len(train_loader.dataset)
        class_loss_avg = class_loss_total / max(source_sample_count, 1)
        train_acc /= max(source_sample_count, 1)

        val_m = eval_model_metrics(model, val_loader, device, n_classes=n_classes)
        test_m = eval_model_metrics(model, test_loader, device, n_classes=n_classes)
        full_m = eval_model_metrics(model, full_test_loader, device, n_classes=n_classes)

        if (best_val_metrics is None) or (val_m["acc"] > best_val_metrics["acc"]):
            best_val_metrics = val_m
            best_test_metrics = test_m
            best_full_metrics = full_m
            torch.save(
                {
                    "backbone": model.backbone.state_dict(),
                    "sequential": model.sequential.state_dict(),
                    "classifier": model.classifier.state_dict(),
                },
                f"{save}/stage3_best.pth",
            )

        stage3_metrics.append({
            "Epoch": epoch + 1,
            "Total_Loss": round(total_loss, 4),
            "Class_Loss": round(class_loss_avg, 4),
            "Train_Acc(source)": round(train_acc, 4),

            "Val_Acc(acc2)": round(val_m["acc"], 4),
            "Val_Kappa": round(val_m["kappa"], 4),
            "Val_F1": round(val_m["f1"], 4),
            "Val_AUC": round(val_m["auc"], 4) if not np.isnan(val_m["auc"]) else val_m["auc"],

            "Test_Acc(acc1)": round(test_m["acc"], 4),
            "Test_Kappa": round(test_m["kappa"], 4),
            "Test_F1": round(test_m["f1"], 4),
            "Test_AUC": round(test_m["auc"], 4) if not np.isnan(test_m["auc"]) else test_m["auc"],

            "Full_Acc(acc3)": round(full_m["acc"], 4),
            "Full_Kappa": round(full_m["kappa"], 4),
            "Full_F1": round(full_m["f1"], 4),
            "Full_AUC": round(full_m["auc"], 4) if not np.isnan(full_m["auc"]) else full_m["auc"],

            "Best_Val_Acc": round(best_val_metrics["acc"], 4) if best_val_metrics else 0.0,
        })

    checkpoint = torch.load(f"{save}/stage3_best.pth", map_location=device)
    model.backbone.load_state_dict(checkpoint["backbone"])
    model.sequential.load_state_dict(checkpoint["sequential"])
    model.classifier.load_state_dict(checkpoint["classifier"])

    return model, best_val_metrics, best_test_metrics, best_full_metrics


def train_target_only(model, train_loader, test_loader, val_loader, full_test_loader, device,
                      epochs=300, lr=5e-4, n_classes=4):
    global target_only_metrics
    model.to(device)

    criterion = nn.NLLLoss()
    optimizer = optim.Adam(model.parameters(), lr=lr)
    constraint = MaxNormDefaultConstraint()

    best_val_metrics = None
    best_test_metrics = None
    best_full_metrics = None

    for epoch in range(epochs):
        model.train()
        train_loss = 0.0
        train_acc = 0.0

        for data, label in train_loader:
            data, label = data.to(device), label.to(device)
            optimizer.zero_grad()
            output = model(data)
            loss = criterion(output, label)
            loss.backward()
            optimizer.step()
            constraint.apply(model)

            bs = data.size(0)
            train_loss += loss.item() * bs
            train_acc += calculate_accuracy(output, label) * bs

        train_loss /= len(train_loader.dataset)
        train_acc /= len(train_loader.dataset)

        test_m = eval_model_metrics(model, test_loader, device, n_classes=n_classes)
        val_m = eval_model_metrics(model, val_loader, device, n_classes=n_classes)
        full_m = eval_model_metrics(model, full_test_loader, device, n_classes=n_classes)

        if (best_val_metrics is None) or (val_m["acc"] > best_val_metrics["acc"]):
            best_val_metrics = val_m
            best_test_metrics = test_m
            best_full_metrics = full_m
            torch.save(model.state_dict(), f"{save}/target_only_best.pth")

        target_only_metrics.append({
            "Epoch": epoch + 1,
            "Train_Loss": round(train_loss, 4),
            "Train_Acc": round(train_acc, 4),

            "Test_Acc(acc1)": round(test_m["acc"], 4),
            "Test_Kappa": round(test_m["kappa"], 4),
            "Test_F1": round(test_m["f1"], 4),
            "Test_AUC": round(test_m["auc"], 4) if not np.isnan(test_m["auc"]) else test_m["auc"],

            "Val_Acc(acc2)": round(val_m["acc"], 4),
            "Val_Kappa": round(val_m["kappa"], 4),
            "Val_F1": round(val_m["f1"], 4),
            "Val_AUC": round(val_m["auc"], 4) if not np.isnan(val_m["auc"]) else val_m["auc"],

            "Full_Acc(acc3)": round(full_m["acc"], 4),
            "Full_Kappa": round(full_m["kappa"], 4),
            "Full_F1": round(full_m["f1"], 4),
            "Full_AUC": round(full_m["auc"], 4) if not np.isnan(full_m["auc"]) else full_m["auc"],

            "Best_Val_Acc": round(best_val_metrics["acc"], 4) if best_val_metrics else 0.0,
        })

    model.load_state_dict(torch.load(f"{save}/target_only_best.pth", map_location=device))
    return best_test_metrics, best_val_metrics, best_full_metrics


def train_source_only(model, train_loader, test_loader, val_loader, full_test_loader, device,
                      epochs=300, lr=5e-4, n_classes=4):
    global source_only_metrics
    model.to(device)

    criterion = nn.NLLLoss()
    optimizer = optim.Adam(model.parameters(), lr=lr)
    constraint = MaxNormDefaultConstraint()

    best_val_metrics = None
    best_test_metrics = None
    best_full_metrics = None

    for epoch in range(epochs):
        model.train()
        train_loss = 0.0
        train_acc = 0.0

        for data, label in train_loader:
            data, label = data.to(device), label.to(device)
            optimizer.zero_grad()
            output = model(data)
            loss = criterion(output, label)
            loss.backward()
            optimizer.step()
            constraint.apply(model)

            bs = data.size(0)
            train_loss += loss.item() * bs
            train_acc += calculate_accuracy(output, label) * bs

        train_loss /= len(train_loader.dataset)
        train_acc /= len(train_loader.dataset)

        test_m = eval_model_metrics(model, test_loader, device, n_classes=n_classes)
        val_m = eval_model_metrics(model, val_loader, device, n_classes=n_classes)
        full_m = eval_model_metrics(model, full_test_loader, device, n_classes=n_classes)

        if (best_val_metrics is None) or (val_m["acc"] > best_val_metrics["acc"]):
            best_val_metrics = val_m
            best_test_metrics = test_m
            best_full_metrics = full_m
            torch.save(model.state_dict(), f"{save}/source_only_best.pth")

        source_only_metrics.append({
            "Epoch": epoch + 1,
            "Train_Loss": round(train_loss, 4),
            "Train_Acc": round(train_acc, 4),

            "Test_Acc(acc1)": round(test_m["acc"], 4),
            "Test_Kappa": round(test_m["kappa"], 4),
            "Test_F1": round(test_m["f1"], 4),
            "Test_AUC": round(test_m["auc"], 4) if not np.isnan(test_m["auc"]) else test_m["auc"],

            "Val_Acc(acc2)": round(val_m["acc"], 4),
            "Val_Kappa": round(val_m["kappa"], 4),
            "Val_F1": round(val_m["f1"], 4),
            "Val_AUC": round(val_m["auc"], 4) if not np.isnan(val_m["auc"]) else val_m["auc"],

            "Full_Acc(acc3)": round(full_m["acc"], 4),
            "Full_Kappa": round(full_m["kappa"], 4),
            "Full_F1": round(full_m["f1"], 4),
            "Full_AUC": round(full_m["auc"], 4) if not np.isnan(full_m["auc"]) else full_m["auc"],

            "Best_Val_Acc": round(best_val_metrics["acc"], 4) if best_val_metrics else 0.0,
        })

    model.load_state_dict(torch.load(f"{save}/source_only_best.pth", map_location=device))
    return best_test_metrics, best_val_metrics, best_full_metrics


def train_source_only_euclid(model, train_loader, test_loader, val_loader, full_test_loader, device,
                             epochs=300, lr=5e-4, n_classes=4):
    global source_only_euclid_metrics
    model.to(device)

    criterion = nn.NLLLoss()
    optimizer = optim.Adam(model.parameters(), lr=lr)
    constraint = MaxNormDefaultConstraint()

    best_val_metrics = None
    best_test_metrics = None
    best_full_metrics = None

    for epoch in range(epochs):
        model.train()
        train_loss = 0.0
        train_acc = 0.0

        for data, label in train_loader:
            data, label = data.to(device), label.to(device)
            optimizer.zero_grad()
            output = model(data)
            loss = criterion(output, label)
            loss.backward()
            optimizer.step()
            constraint.apply(model)

            bs = data.size(0)
            train_loss += loss.item() * bs
            train_acc += calculate_accuracy(output, label) * bs

        train_loss /= len(train_loader.dataset)
        train_acc /= len(train_loader.dataset)

        test_m = eval_model_metrics(model, test_loader, device, n_classes=n_classes)
        val_m = eval_model_metrics(model, val_loader, device, n_classes=n_classes)
        full_m = eval_model_metrics(model, full_test_loader, device, n_classes=n_classes)

        if (best_val_metrics is None) or (val_m["acc"] > best_val_metrics["acc"]):
            best_val_metrics = val_m
            best_test_metrics = test_m
            best_full_metrics = full_m
            torch.save(model.state_dict(), f"{save}/source_only_euclid_best.pth")

        source_only_euclid_metrics.append({
            "Epoch": epoch + 1,
            "Train_Loss": round(train_loss, 4),
            "Train_Acc": round(train_acc, 4),

            "Test_Acc(acc1)": round(test_m["acc"], 4),
            "Test_Kappa": round(test_m["kappa"], 4),
            "Test_F1": round(test_m["f1"], 4),
            "Test_AUC": round(test_m["auc"], 4) if not np.isnan(test_m["auc"]) else test_m["auc"],

            "Val_Acc(acc2)": round(val_m["acc"], 4),
            "Val_Kappa": round(val_m["kappa"], 4),
            "Val_F1": round(val_m["f1"], 4),
            "Val_AUC": round(val_m["auc"], 4) if not np.isnan(val_m["auc"]) else val_m["auc"],

            "Full_Acc(acc3)": round(full_m["acc"], 4),
            "Full_Kappa": round(full_m["kappa"], 4),
            "Full_F1": round(full_m["f1"], 4),
            "Full_AUC": round(full_m["auc"], 4) if not np.isnan(full_m["auc"]) else full_m["auc"],

            "Best_Val_Acc": round(best_val_metrics["acc"], 4) if best_val_metrics else 0.0,
        })

    model.load_state_dict(torch.load(f"{save}/source_only_euclid_best.pth", map_location=device))
    return best_test_metrics, best_val_metrics, best_full_metrics


# ===================== 保存 Excel（原样） =====================
def save_all_results_to_excel(output_path, lambda2,
                              stage1_res, stage2_res, stage3_res,
                              target_only_res, source_only_res, source_only_euclid_res):
    wb = Workbook()
    wb.remove(wb.active)

    def _write_df(sheet_name, rows_list):
        if not rows_list:
            return
        ws = wb.create_sheet(title=sheet_name)
        df = pd.DataFrame(rows_list)
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            max_len = max(len(str(cell.value)) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 28)

    _write_df("Stage1_Unsupervised", stage1_metrics)
    _write_df("Stage2_Discriminators", stage2_metrics)
    _write_df("Stage3_Adversarial", stage3_metrics)
    _write_df("Target_Only (target train)", target_only_metrics)
    _write_df("Source_Only (source train)", source_only_metrics)
    _write_df("Source_Only_EuclidAlign", source_only_euclid_metrics)

    ws = wb.create_sheet(title="All_Methods_Summary")

    def _pick(m: dict, key: str):
        v = m.get(key, float("nan"))
        if isinstance(v, float) and np.isnan(v):
            return v
        return round(v, 4) if isinstance(v, (float, int)) else v

    ours_val, ours_test, ours_full = stage3_res
    t_test, t_val, t_full = target_only_res
    s_test, s_val, s_full = source_only_res
    se_test, se_val, se_full = source_only_euclid_res

    summary_rows = [
        {
            "Method": "Ours(Stage3)",
            "acc1_acc": _pick(ours_test, "acc"), "acc1_kappa": _pick(ours_test, "kappa"), "acc1_f1": _pick(ours_test, "f1"), "acc1_auc": _pick(ours_test, "auc"),
            "acc2_acc": _pick(ours_val, "acc"),  "acc2_kappa": _pick(ours_val, "kappa"),  "acc2_f1": _pick(ours_val, "f1"),  "acc2_auc": _pick(ours_val, "auc"),
            "acc3_acc": _pick(ours_full, "acc"), "acc3_kappa": _pick(ours_full, "kappa"), "acc3_f1": _pick(ours_full, "f1"), "acc3_auc": _pick(ours_full, "auc"),
            "Lambda2": lambda2,
            "Target": f"sub{TARGET_SUBJECT}",
            "ShotsPerClass": SAMPLES_PER_CLASS,
            "ShotsTotal": SAMPLES_PER_CLASS * N_CLASSES,
        },
        {
            "Method": "TargetOnly(target train)",
            "acc1_acc": _pick(t_test, "acc"), "acc1_kappa": _pick(t_test, "kappa"), "acc1_f1": _pick(t_test, "f1"), "acc1_auc": _pick(t_test, "auc"),
            "acc2_acc": _pick(t_val, "acc"),  "acc2_kappa": _pick(t_val, "kappa"),  "acc2_f1": _pick(t_val, "f1"),  "acc2_auc": _pick(t_val, "auc"),
            "acc3_acc": _pick(t_full, "acc"), "acc3_kappa": _pick(t_full, "kappa"), "acc3_f1": _pick(t_full, "f1"), "acc3_auc": _pick(t_full, "auc"),
            "Lambda2": "-",
            "Target": f"sub{TARGET_SUBJECT}",
            "ShotsPerClass": SAMPLES_PER_CLASS,
            "ShotsTotal": SAMPLES_PER_CLASS * N_CLASSES,
        },
        {
            "Method": "SourceOnly(source train)",
            "acc1_acc": _pick(s_test, "acc"), "acc1_kappa": _pick(s_test, "kappa"), "acc1_f1": _pick(s_test, "f1"), "acc1_auc": _pick(s_test, "auc"),
            "acc2_acc": _pick(s_val, "acc"),  "acc2_kappa": _pick(s_val, "kappa"),  "acc2_f1": _pick(s_val, "f1"),  "acc2_auc": _pick(s_val, "auc"),
            "acc3_acc": _pick(s_full, "acc"), "acc3_kappa": _pick(s_full, "kappa"), "acc3_f1": _pick(s_full, "f1"), "acc3_auc": _pick(s_full, "auc"),
            "Lambda2": "-",
            "Target": f"sub{TARGET_SUBJECT}",
            "ShotsPerClass": SAMPLES_PER_CLASS,
            "ShotsTotal": SAMPLES_PER_CLASS * N_CLASSES,
        },
        {
            "Method": "SourceOnly+EuclidAlign",
            "acc1_acc": _pick(se_test, "acc"), "acc1_kappa": _pick(se_test, "kappa"), "acc1_f1": _pick(se_test, "f1"), "acc1_auc": _pick(se_test, "auc"),
            "acc2_acc": _pick(se_val, "acc"),  "acc2_kappa": _pick(se_val, "kappa"),  "acc2_f1": _pick(se_val, "f1"),  "acc2_auc": _pick(se_val, "auc"),
            "acc3_acc": _pick(se_full, "acc"), "acc3_kappa": _pick(se_full, "kappa"), "acc3_f1": _pick(se_full, "f1"), "acc3_auc": _pick(se_full, "auc"),
            "Lambda2": "-",
            "Target": f"sub{TARGET_SUBJECT}",
            "ShotsPerClass": SAMPLES_PER_CLASS,
            "ShotsTotal": SAMPLES_PER_CLASS * N_CLASSES,
        },
    ]

    df_sum = pd.DataFrame(summary_rows)
    for r in dataframe_to_rows(df_sum, index=False, header=True):
        ws.append(r)

    for cell in ws[1]:
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal="center")

    for col in ws.columns:
        max_len = max(len(str(cell.value)) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    wb.save(output_path)


# ===================== 单次实验：给定 target_subject & total_shots =====================
def run_one_experiment(target_subject: int, total_shots: int):
    global TARGET_SUBJECT, SAMPLES_PER_CLASS, save

    samples_per_class = total_shots // N_CLASSES
    actual_total = samples_per_class * N_CLASSES

    TARGET_SUBJECT = target_subject
    SAMPLES_PER_CLASS = samples_per_class

    save = os.path.join(SAVE_ROOT, f"sub{target_subject}", str(total_shots))
    check_and_create_folder(save)

    np.random.seed(RANDOM_SEED)
    torch.manual_seed(RANDOM_SEED)
    if torch.cuda.is_available():
        torch.cuda.manual_seed(RANDOM_SEED)
        torch.cuda.manual_seed_all(RANDOM_SEED)

    logging.basicConfig(format="%(asctime)s %(levelname)s : %(message)s", level=logging.INFO, stream=sys.stdout)
    logging.info(f"========== RUN: target=sub{target_subject}, folderShots={total_shots}, perClass={samples_per_class}, actualTotal={actual_total} ==========")

    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    reset_global_metrics()

    (
        train_loader, stage2_train_loader, target_train_loader, source_train_loader,
        target_shot_loader, target_val_loader, target_full_test_loader,
        source_train_loader_euclid, target_shot_loader_euclid, target_val_loader_euclid, target_full_test_loader_euclid,
        _, _, _
    ) = load_all_data(DATA_FOLDER, target_subject, BATCH_SIZE, samples_per_class)

    # ===================== 这里把模型换成 ADFCNNNetWrapper，其它逻辑不动 =====================
    n_ch = 22
    in_t = 1125  # ival[-500,4000]@250Hz => 4.5s => 1125点

    model_ours = ADFCNNNetWrapper(n_classes=N_CLASSES, in_chans=n_ch, in_samples=in_t, return_fc=False)
    model_target_only = ADFCNNNetWrapper(n_classes=N_CLASSES, in_chans=n_ch, in_samples=in_t, return_fc=False)
    model_source_only = ADFCNNNetWrapper(n_classes=N_CLASSES, in_chans=n_ch, in_samples=in_t, return_fc=False)
    model_source_only_euclid = ADFCNNNetWrapper(n_classes=N_CLASSES, in_chans=n_ch, in_samples=in_t, return_fc=False)
    # ======================================================================================

    # 判别器输入维度（保持你原逻辑：features = sequential(backbone(x))）
    model_ours.to(device)
    with torch.no_grad():
        dummy_input = torch.randn(1, n_ch, in_t, 1).to(device)
        features = model_ours.sequential(model_ours.backbone(dummy_input))
        final_ch = features.shape[1]
        final_conv_length = features.shape[2]

    domain_disc = create_domain_discriminator(in_channels=final_ch, in_length=final_conv_length)
    subject_disc = create_subject_discriminator(in_channels=final_ch, in_length=final_conv_length)

    # Stage1
    model_ours, stage1_val_m, stage1_test_m, stage1_full_m = stage1_unsupervised_train(
        model_ours, train_loader, target_val_loader, target_shot_loader, target_full_test_loader,
        device, epochs=EPOCHS, lr=LR, n_classes=N_CLASSES
    )

    # Stage2
    domain_disc, subject_disc, stage2_test_domain, stage2_test_subject, stage2_full_domain, stage2_full_subject = stage2_train_discriminators(
        model_ours.backbone, model_ours.sequential, domain_disc, subject_disc,
        stage2_train_loader, target_val_loader, target_shot_loader, target_full_test_loader,
        device, target_subject_idx=target_subject - 1, epochs=EPOCHS, lr=LR
    )

    # Stage3
    model_ours, stage3_val_m, stage3_test_m, stage3_full_m = stage3_adversarial_train(
        model_ours, domain_disc, subject_disc,
        train_loader, target_val_loader, target_shot_loader, target_full_test_loader,
        device, epochs=EPOCHS, lr=LR, lambda1=LAMBDA1, lambda2=LAMBDA2, n_classes=N_CLASSES
    )

    # Baseline1
    target_only_test_m, target_only_val_m, target_only_full_m = train_target_only(
        model_target_only, target_train_loader, target_shot_loader, target_val_loader, target_full_test_loader,
        device, epochs=EPOCHS, lr=LR, n_classes=N_CLASSES
    )

    # Baseline2
    source_only_test_m, source_only_val_m, source_only_full_m = train_source_only(
        model_source_only, source_train_loader, target_shot_loader, target_val_loader, target_full_test_loader,
        device, epochs=EPOCHS, lr=LR, n_classes=N_CLASSES
    )

    # Baseline3
    source_euclid_test_m, source_euclid_val_m, source_euclid_full_m = train_source_only_euclid(
        model_source_only_euclid, source_train_loader_euclid,
        target_shot_loader_euclid, target_val_loader_euclid, target_full_test_loader_euclid,
        device, epochs=EPOCHS, lr=LR, n_classes=N_CLASSES
    )

    output_path = os.path.join(save, f"sub{target_subject}_all_methods_results.xlsx")
    save_all_results_to_excel(
        output_path, LAMBDA2,
        stage1_res=(stage1_val_m, stage1_test_m, stage1_full_m),
        stage2_res=(stage2_test_domain, stage2_test_subject, stage2_full_domain, stage2_full_subject),
        stage3_res=(stage3_val_m, stage3_test_m, stage3_full_m),
        target_only_res=(target_only_test_m, target_only_val_m, target_only_full_m),
        source_only_res=(source_only_test_m, source_only_val_m, source_only_full_m),
        source_only_euclid_res=(source_euclid_test_m, source_euclid_val_m, source_euclid_full_m),
    )

    del model_ours, model_target_only, model_source_only, model_source_only_euclid, domain_disc, subject_disc
    if torch.cuda.is_available():
        torch.cuda.empty_cache()

    logging.info(f"Saved to: {save}")


# ===================== 主入口：sub1~sub9 × shotsList =====================
def main():
    check_and_create_folder(SAVE_ROOT)

    for sub in range(1, 10):
        for total in SHOT_TOTAL_LIST:
            run_one_experiment(target_subject=sub, total_shots=total)

    print("\n全部实验完成。")


if __name__ == "__main__":
    main()
